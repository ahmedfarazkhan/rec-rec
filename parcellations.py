# -*- coding: utf-8 -*-
"""
Map data across parcellations

@author: akhan


Data 
- AtlasesMRI/brodmann.hdr
- AtlasesMRI/brodmann.img
- AtlasesMRI/Julich_Anatomy_v22c.nii
- AtlasesMRI/Julich_Anatomy_v22c_MPM.mat

- Human/receptor_data.txt
- Human/receptor_data.xls
- Human/receptor_data_area_atlas_correlation.xlsx

"""

import os
import numpy as np
import h5py
from openpyxl import load_workbook
import csv

N_SUBJECTS = 20 #200
N_REGIONS = 40
N_FACTORS = 6
N_RECEPTORS = 19

def get_MAT_data(field_name, index):
    reference = f_julich_metadata["MAP"][field_name][index][0]
    return f_julich_metadata[reference] 
    
# Match areas  
path_julich_metadata = os.path.join(os.getcwd(), 'data/AtlasesMRI/Julich_Anatomy_v22c_MPM.mat')
path_human_metadata = os.path.join(os.getcwd(), 'data/Human/receptor_data_area_atlas_correlation.xlsx')
path_julich_rec_data = os.path.join(os.getcwd(), 'data/Human/receptor_data.txt')


###############################################################################

f_julich_metadata = h5py.File(path_julich_metadata, 'r')
julich_map = f_julich_metadata['MAP']

julich_names = f_julich_metadata['MAP']['name']

# List all fields
print("Fields: %s" %f_julich_metadata["MAP"].keys())

field_name = 'name'
julich_area_names = []
for i in range(f_julich_metadata['MAP'][field_name].shape[0]):
    area_name_arr = np.asarray(get_MAT_data(field_name=field_name, index=i)[...])
    julich_area_names += [''.join([chr(c_arr[0]) for c_arr in area_name_arr])]

print julich_area_names

###############################################################################
print "\n"

# Copy spreadsheet data
wb = load_workbook(filename=path_human_metadata)
ws = wb[wb.sheetnames[0]]

# [Region, Julich brain atlas]
area_atlas = []

# Copy data from spreadsheet
for row in ws.rows:#[:1]: 
    r = []
    for cell in row[:2]:
        # Long to int
        if type(cell.value) == type(0L):
            r += [int(cell.value)]
        # Unicode to ascii
        elif type(cell.value) == type(u'unicode'):
            r += [(cell.value).encode('ascii', 'ignore')]
    area_atlas += [r]

area_atlas = np.asarray(area_atlas)
    
# Last col is empty
#area_atlas = np.delete(area_atlas, (2), axis=1)

# Convert to dictionary
julich_area_conversion = dict(zip(list(area_atlas[:,0]), list(area_atlas[:,1])))

##############################################################################
julich_regions = []

with open(path_julich_rec_data, 'rb') as f_rec_data:
    recdata = csv.reader(f_rec_data, delimiter=' ', quotechar='|')
    for row in recdata:
        julich_regions += [row[0]]
        
##############################################################################

# Compare area_atlas to julich_regions
brain_regions = []

for region in julich_regions:
    brain_regions += [julich_area_conversion[region]]




        