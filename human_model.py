# -*- coding: utf-8 -*-
"""
Human receptor density model

Data:

- Human/receptor_data.txt
- Human/receptor_data.xls
- Human/receptor_data_area_atlas_correlation.xlsx

@author: akhan
"""

import csv
import pandas as pd
import numpy as np
from openpyxl import load_workbook

path_human_densities = os.path.join(os.getcwd(), 'data/Human/receptor_data.txt')
path_mice_densities = os.path.join(os.getcwd(), 'data/Knockout mice_data/densities_5HT1A_KO_model_control_KO_animals.xlsx')

data = pd.read_csv(path_human_densities, sep=" ", header=None)

human_receptors = data.loc[0].as_matrix()[1:]


# Load mouse receptors
wb = load_workbook(filename=path_mice_densities)
ws = wb[wb.sheetnames[0]]

rec_densities = []
for row in ws.rows:
    r = []
    for cell in row:
        r += [cell.value]
    rec_densities += [r]
 
mouse_receptors = np.asarray([str(r) for r in rec_densities[0][3:-2]] )

common_receptors = ['AMPA', 'MK80?', 'KAIN', 'MUSC', 'FLUM', 'CGP5?', 'PIRE', 'OXOT', 'DAMP', 'EPIB', 'PRAZ', 'KETA', 'DPAT', 'SCH?' ]
missing_mouse_receptors = ['LY34', 'UK14', 'RACL', 'DPMG'] # RACL only for basal ganglia
missing_human_receptors = ['RX']


'''
Mouse receptors
AMPA  [3H] AMPA  AMPA receptor (glutamate)
MK80  [3H] MK801  NMDA receptor  (glutamate)
KAIN  [3H] kainate  kainate receptor (glutamate)
LY34  [3H] LY 341 495  mGluR2/3 receptor (glutamate)
MUSC  [3H] muscimol  GABAa receptor (GABA)
FLUM  [3H] flumazenil  GABAa associaated benzodiazepine binding site (GABA)
CGP5  [3H] CGP54626  GABAb receptor (GABA)
PIRE  [3H] pirenzepine  muscarinic M1 receptor (acetylcholine)
OXOT  [3H] oxotremorine-M  muscarinic M2 receptor (acetylcholine)
DAMP  [3H] 4-DAMP  muscarinic M3 receptor (acetylcholine)
EPIB  [3H] epibatidine  nicotinic alpha4/beta2 receptor (acetylcholine)
PRAZ  [3H] prazosin  alpha1 receptor (noradrenlalin)
UK14  [3H] UK14-304  alpha2 receptor (noradrenlalin)
KETA  [3H] ketanserin  5-HT2 receptor (serotonin)
DPAT  [3H] 8-OH-DPAT  5-HT1A receptor (serotonin)
SCH2  [3H] SCH23390  D1 receptor (dopamine)
RACL  [3H] raclopride  D2 receptor (dopamine) (only data for basal ganglia)
DPMG  [3H] DPMG  A1 receptor (adenosine)
'''