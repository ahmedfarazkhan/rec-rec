# -*- coding: utf-8 -*-
"""
Receptor interactions in mice

Model the dynamic system (using the control human and mice data), and analyze 
if we can make predictions of interventions, validating with the knockout data.
Nomenclature is from the 2001 Paxinos atlas of the mouse brain


@author: akhan

Step 1: Adjacency matrix A
Step 2: Predict intervention time, predict final intervention state
Step 3: Compare predictions with gene expression

Data:

- Knockout mice_data/densities_5HT1A_KO_model_control_KO_animals.xlsx
- Human/A_recp_vs_recp_matrix_human.mat


for this Wednesday, could you please construct a matrix in which each element i,j is 
the MECS value for the region i and the receptor j (for the mice data). For t0 and tf,
 just use something like 0 and 1 (and do not enter any value in driver_nodes and times_step, 
 it will be set automatically).
"""

import os
import numpy as np
import scipy.io
import pandas as pd
import h5py

from openpyxl import load_workbook

N_META_COLS = 3 # Metadata in Excel file (group, animal, region)

N_MICE_KO = 6
N_MICE_CTRL = 7
N_RECEPTORS = 18
N_REGIONS = 20

missing_mice_indices = [3, 16, 17] # ['LY34',  'RACL', 'DPMG']
mice_rec_in_human_indices = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 13, 12, 14] 

path_mice_densities = os.path.join(os.getcwd(), 'data/Knockout mice_data/densities_5HT1A_KO_model_control_KO_animals.xlsx')
path_human_A = os.path.join(os.getcwd(), 'data/Human/A_recp_vs_recp_matrix_human.mat')

def replace_none(data, value=np.nan):
    for mouse in range(len(data)):
        for rec_reg in range(data.shape[-1]):
            if data[mouse, rec_reg] == None:
                data[mouse, rec_reg] = np.nan
    return data

def copy_triangular(matrix):
    n_rows, n_cols = matrix.shape
    for i in range(n_rows):
        for j in range(i, n_cols):
            matrix[j][i] = matrix[i][j]
    return matrix

def row_normalize(matrix):
    #row_sums = matrix.sum(axis=1)
    row_maxes = np.amax(matrix, axis=1)
    return matrix / row_maxes[:, np.newaxis]
    
# Copy spreadsheet data
wb = load_workbook(filename=path_mice_densities)
ws = wb[wb.sheetnames[0]]

# [group, animal, region, rec1, rec2, ..., recN]

rec_densities = []

# Copy data from spreadsheet
for row in ws.rows:
    r = []
    for cell in row:
        r += [cell.value]
    rec_densities += [r]
    
# Prune empty cells
rec_densities = np.asarray(rec_densities)[: (N_MICE_KO + N_MICE_CTRL)*N_REGIONS + 1, : N_RECEPTORS + N_META_COLS]

# Remove missing receptors and rearrange
#rec_data = np.delete(rec_densities, (N_META_COLS+N_RECEPTORS-2), axis=1)
for missing_index in reversed(missing_mice_indices):
    print "Deleting receptor index",  missing_index
    rec_densities = np.delete(rec_densities, (N_META_COLS + missing_index), axis=1)
    N_RECEPTORS -= 1
    
# No metadata and rearranged
rec_densities = rec_densities[1:, N_META_COLS:]
rec_densities = rec_densities[:, mice_rec_in_human_indices]
    
# No metadata
rec_ctrl = rec_densities[:N_REGIONS*N_MICE_CTRL,:]
rec_ko = rec_densities[N_REGIONS*N_MICE_CTRL:,:]

# Add another dimension for animals
rec_ctrl = rec_ctrl.reshape((N_MICE_CTRL, N_REGIONS, N_RECEPTORS))
rec_ko = rec_ko.reshape((N_MICE_KO, N_REGIONS, N_RECEPTORS))

# Reshape for interpolation processing
ctrl_2d = rec_ctrl.reshape((N_MICE_CTRL, N_REGIONS * N_RECEPTORS))
ko_2d = rec_ko.reshape((N_MICE_KO, N_REGIONS * N_RECEPTORS))


# Impute missing data
ctrl_2d = replace_none(ctrl_2d)
ko_2d = replace_none(ko_2d)

#ctrl_df = pd.DataFrame(data=ctrl_2d)
#ko_df = pd.DataFrame(data=ko_2d)

#print(ctrl_df.isnull().sum())
#print(ko_df.isnull().sum())

# Todo: use different imputation
#ctrl_df.fillna(ctrl_df.mean(), inplace=True)
#ko_df.fillna(ko_df.mean(), inplace=True)

# Todo: Normalize
#ctrl_3d = ctrl_2d.reshape((N_MICE_CTRL, N_REGIONS, N_RECEPTORS))
#ko_3d = ko_2d.reshape((N_MICE_KO, N_REGIONS, N_RECEPTORS))

# Save as Matlab matrices
scipy.io.savemat(os.path.join(os.getcwd(), 'data/Knockout mice_data/ctrl_densities.mat'), mdict={'rec_ctrl': ctrl_2d.astype('double')})
scipy.io.savemat(os.path.join(os.getcwd(), 'data/Knockout mice_data/ko_densities.mat'), mdict={'rec_ko': ko_2d.astype('double')})

'''

## Todo: Interpolation of missing receptor densities

# TODO: S equilibrium states from receptor densities?
# or is it propagation weight parameters
S = np.random.rand(N_RECEPTORS)

# TODO: build MCM matrix A
# A_m,n(i,j;t) 
# m,n index factors
# i,j index regions

# Todo: replace with real mouse brain connectivity data
# Placeholder for inter-factor influence a^n->m
a = copy_triangular(np.random.rand((N_RECEPTORS, N_RECEPTORS))) # Symmetric
# Placeholder for anatomical and vascular connectivity C^m_i->j
C = copy_triangular(np.random.rand((N_REGIONS, N_REGIONS))) # Symmetric

# Todo: Convert to brain state s (N_REGIONS * N_RECEPTORS)

A = np.zeros((N_REGIONS*N_RECEPTORS, N_REGIONS*N_RECEPTORS))

for row in range(N_REGIONS*N_RECEPTORS):
    
    rec_from = row % N_REGIONS
    reg_from = row % N_RECEPTORS
    
    for col in range(N_REGIONS*N_RECEPTORS):
        
        rec_to = col % N_REGIONS
        reg_to = col % N_RECEPTORS
        
        # Diagonals
        if (rec_from == rec_to) and (reg_from == reg_to):
            net_flux = C[reg_from, :] # TODO *S
            A[row, col] = a[rec_from, rec_to] - net_flux
        # Local inter-factor
        elif (rec_from != rec_to) and (reg_from == reg_to):
            A[row, col] = a[rec_from, rec_to] # TODO: check index order
        # Inter-region 
        elif (rec_from == rec_to) and (reg_from != reg_to):
            A[row, col] = C[reg_from, reg_to] # TODO * S
        # No inter-region inter-factorial influence
        elif (rec_from != rec_to) and (reg_from != reg_to):
            A[row, col] = 0
'''